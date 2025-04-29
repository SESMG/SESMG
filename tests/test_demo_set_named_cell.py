import pytest
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

import sys
import os
import importlib.util

# Absolute path to the file 4_Demo_Tool.py
tool_path = os.path.abspath(
    os.path.join(os.path.dirname(__file__), '..', 'program_files', 'GUI_st', 'pages', '4_Demo_Tool.py')
)

# Load module dynamically using importlib
spec = importlib.util.spec_from_file_location("demo_tool", tool_path)
demo_tool = importlib.util.module_from_spec(spec)
spec.loader.exec_module(demo_tool)

def test_set_named_cell_existing_name():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = ""

    dn = DefinedName(name="MyCell", attr_text="Sheet1!$A$1")
    wb.defined_names.add(dn)  # use .add() instead of .append()

    demo_tool.set_named_cell(wb, "MyCell", 123)

    assert ws["A1"].value == 123

def test_set_named_cell_non_existing_name(capfd):
    wb = Workbook()

    demo_tool.set_named_cell(wb, "UnknownName", 999)

    out, err = capfd.readouterr()
    assert "Error: 'UnknownName' is not defined in the Excel file." in out